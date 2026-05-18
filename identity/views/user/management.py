import logging
import itertools
from datetime import timedelta
import re

from django.core.cache import cache
from django.core.exceptions import PermissionDenied
from django.db.models import Q, QuerySet
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.conf import settings
from django.utils.decorators import method_decorator
from django.core.files.uploadedfile import UploadedFile

from rest_framework.settings import api_settings
from rest_framework.generics import (
    CreateAPIView,
    ListAPIView,
    RetrieveUpdateDestroyAPIView,
)
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.exceptions import ValidationError

from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi

from vmlc.serializers import (
    StaffListSerializer,
    StaffInviteSerializer,
    CandidateListSerializer,
    UserProfileDetailSerializer,
    UserProfileListSerializer,
)
from identity.models import PreRegUser, User, UserVerification, Staff, Candidate
from vmlc.utils.auth import generate_password
from vmlc.utils.swagger_schemas import (
    api_key,
    bearer_auth,
    staff_registration_request_body,
    candidate_registration_request_body,
    account_management_response_schema,
    error_response_401,
    error_response_403,
    error_response_404,
    error_response_400,
)
from identity.permissions import (
    AuthenticatedUser,
    IsManagerForStaffDetail,
    IsObjectOwnerOrActiveAdmin,
    ActiveAdminPermissions,
    ActiveManagerPermissions,
    ActiveModeratorPermissions,
)
from vmlc.tasks import (
    revoke_user_invite_task,
)
from comms.tasks import send_mail_task
from vmlc.utils.stats import generate_stats_overview_data
from vmlc.utils.query_filters import (
    filter_pre_reg_users,
    filter_staffs,
    filter_candidates,
    filter_users,
)
from vmlc.v2.serializers.registration import PreRegUserSerializer
from vmlc.v2.utils import get_or_set_cache, CacheKeys

logger = logging.getLogger(__name__)


class RequestDataExtractor:
    """Extracts and validates user and profile data from request."""

    USER_FIELDS = {"first_name", "last_name", "profile_picture", "phone", "state"}
    PROFILE_FIELDS = {"occupation", "current_class", "school_type"}
    FILE_FIELDS = {"profile_picture"}
    NESTED_PROFILE_PREFIXES = {
        "cowrywise_kid_profile"
    }  # Keys that indicate nested profile data

    @classmethod
    def extract(cls, request_data):
        """
        Extract user and profile data from request, filtering out empty values.
        Returns: (user_data, profile_data)
        """
        user_data = {}
        profile_data = {}

        for key, value in request_data.items():
            if not cls._is_valid_value(key, value):
                continue

            parsed_key = cls._parse_key(key)  # New method for parsing

            if isinstance(
                parsed_key, tuple
            ):  # This is a nested field (e.g., ("cowrywise_kid_profile", "username"))
                parent_key, child_key = parsed_key
                if parent_key in cls.NESTED_PROFILE_PREFIXES:
                    profile_data.setdefault(parent_key, {})[child_key] = value
            elif parsed_key in cls.USER_FIELDS:
                user_data[parsed_key] = value
            elif parsed_key in cls.PROFILE_FIELDS:
                profile_data[parsed_key] = value
            # File fields like 'profile_picture' are handled by _is_valid_value for type checking,
            # but their actual placement into user_data happens here.

        return user_data, profile_data

    @classmethod
    def _is_valid_value(cls, key, value):
        """Check if a value is valid (not empty or placeholder)."""
        if value is None or value == "":
            return False
        if isinstance(value, (list, tuple)) and len(value) == 0:
            return False

        # For file fields, ensure it's an actual uploaded file
        # We need to consider both flat and nested file fields.
        # For now, let's just check if the top-level part of the key is a file field.
        top_level_key = cls._parse_key(key)
        if isinstance(top_level_key, tuple):
            top_level_key = top_level_key[
                0
            ]  # For nested like 'cowrywise_kid_profile.username' -> 'cowrywise_kid_profile'

        if top_level_key in cls.FILE_FIELDS:
            return isinstance(value, UploadedFile)

        return True

    @staticmethod
    def _parse_key(key):
        """
        Parses the incoming key to determine if it's a direct field or a nested field.
        Returns a string for direct fields or a tuple (parent_key, child_key) for nested fields.
        Handles "user.field", "profile.field", "field", "nested.field", "parent[child]".
        """
        logger.debug(f"RequestDataExtractor._parse_key: Processing raw key: {key}")
        original_key = key
        # Handle "user.field_name" or "user[field_name]"
        if key.startswith("user."):
            key = key[5:]  # Remove "user."
        elif key.startswith("user["):
            match = re.match(r"user\[(.*?)\](.*)", key)
            if match:
                key = match.groups()[0] + match.groups()[1]

        # Handle "profile.field_name" or "profile[field_name]"
        # Use original_key for checking "profile." or "profile[" to ensure it's not double processed if 'user' prefix was found
        if original_key.startswith("profile."):
            key = original_key[8:]  # Remove "profile."
        elif original_key.startswith("profile["):
            match = re.match(r"profile\[(.*?)\](.*)", original_key)
            if match:
                key = match.groups()[0] + match.groups()[1]

        # Now, handle nested fields like "cowrywise_kid_profile.username" or "cowrywise_kid_profile[username]"
        if "." in key:
            parts = key.split(".", 1)
            if len(parts) == 2:
                logger.debug(
                    f"RequestDataExtractor._parse_key: Returning tuple: {(parts[0], parts[1])}"
                )
                return (parts[0], parts[1])

        match = re.match(r"(\w+)\[(\w+)\]", key)
        if match:
            logger.debug(
                f"RequestDataExtractor._parse_key: Returning tuple from bracket match: {match.groups()}"
            )
            return match.groups()

        logger.debug(f"RequestDataExtractor._parse_key: Returning direct key: {key}")
        return key


class ProfileManager:
    """Manages profile retrieval and serialization."""

    PROFILE_MAPPING = {
        "candidate_profile": "CandidateDetailSerializer",
        "staff_profile": "StaffDetailSerializer",
    }

    @classmethod
    def get_profile_and_serializer(cls, user):
        """
        Get profile instance and serializer class for user.
        Returns: (profile_instance, serializer_class) or (None, None)
        """
        if hasattr(user, "candidate_profile"):
            from vmlc.serializers import CandidateDetailSerializer

            return user.candidate_profile, CandidateDetailSerializer

        if hasattr(user, "staff_profile"):
            from vmlc.serializers import StaffDetailSerializer

            return user.staff_profile, StaffDetailSerializer

        return None, None

    @classmethod
    def serialize_profile(cls, user):
        """Serialize user profile, removing unnecessary fields."""
        profile, serializer_class = cls.get_profile_and_serializer(user)

        if not profile or not serializer_class:
            return None

        profile_data = serializer_class(profile).data
        profile_data.pop("records", None)

        # Add additional properties
        profile_data["is_setup_complete"] = user.is_setup_complete
        if hasattr(user, "candidate_profile"):
            profile_data["has_cowrywise_kid_profile"] = hasattr(
                user.candidate_profile, "cowrywise_kid_profile"
            )
        else:
            profile_data["has_cowrywise_kid_profile"] = False

        return profile_data


class AccountCacheManager:
    """Handles caching logic for account data."""

    CACHE_TTL = 86400  # 24 hours

    @staticmethod
    def get_cache_key(user_id):
        return f"account_management_{user_id}"

    @classmethod
    def get_cached_data(cls, user_id):
        """Retrieve cached account data."""
        return cache.get(cls.get_cache_key(user_id))

    @classmethod
    def cache_data(cls, user_id, data):
        """Cache account data."""
        cache.set(cls.get_cache_key(user_id), data, cls.CACHE_TTL)

    @classmethod
    def invalidate_user_cache(cls, user):
        """Invalidate all caches related to a user."""
        from identity.models import Staff

        cache.delete(cls.get_cache_key(user.id))

        if hasattr(user, "candidate_profile"):
            cache.delete(f"candidate_dashboard_{user.candidate_profile.pk}")
            from vmlc.utils.helpers import invalidate_all_staff_dashboards

            invalidate_all_staff_dashboards()

        if hasattr(user, "staff_profile"):
            for staff in Staff.objects.all():
                cache.delete(f"staff_dashboard_data_{staff.pk}")


class AccountManagementView(APIView):
    """
    Retrieve or update user account and profile information.

    - GET: Retrieve account and profile information.
    - PATCH: Update account and profile.

    Regular users can manage their own accounts.
    Staff with 'admin' or 'superadmin' roles can manage other users' accounts.
    """

    permission_classes = AuthenticatedUser
    parser_classes = [MultiPartParser, FormParser]

    def _get_target_user(self, request, user_id=None):
        """Determine target user and verify permissions."""
        if user_id is None or user_id == str(request.user.id):
            return request.user

        target_user = get_object_or_404(User, id=user_id)

        if not IsObjectOwnerOrActiveAdmin().has_object_permission(
            request, self, target_user
        ):
            logger.warning(
                f"User {request.user.id} lacks permission to manage user {user_id}"
            )
            raise PermissionDenied("You are not authorized to manage this user.")

        return target_user

    @swagger_auto_schema(
        operation_summary="Get User Account",
        operation_description="Retrieve the account and profile data of the target user.",
        responses={
            200: account_management_response_schema,
            401: error_response_401,
            403: error_response_403,
            404: error_response_404,
        },
        tags=["Account Management"],
        manual_parameters=[api_key, bearer_auth],
    )
    def get(self, request, user_id=None):
        """Retrieve account and profile data."""
        logger.info(
            f"GET account_management: user {request.user.id} requesting user {user_id}"
        )

        target_user = self._get_target_user(request, user_id)

        # Check cache first
        cached_data = AccountCacheManager.get_cached_data(target_user.id)
        if cached_data:
            return Response(cached_data)

        # Serialize profile
        profile_data = ProfileManager.serialize_profile(target_user)
        if not profile_data:
            logger.error(f"User {target_user.id} has no profile")
            return Response(
                {"detail": "User does not have a profile."},
                status=status.HTTP_404_NOT_FOUND,
            )

        response_data = {"profile": profile_data}
        AccountCacheManager.cache_data(target_user.id, response_data)

        return Response(response_data)

    @swagger_auto_schema(
        operation_summary="Update User Account",
        operation_description="Partially update user and/or profile data.",
        manual_parameters=[
            openapi.Parameter(
                "first_name",
                openapi.IN_FORM,
                type=openapi.TYPE_STRING,
                description="User's first name.",
            ),
            openapi.Parameter(
                "last_name",
                openapi.IN_FORM,
                type=openapi.TYPE_STRING,
                description="User's last name.",
            ),
            openapi.Parameter(
                "profile_picture",
                openapi.IN_FORM,
                type=openapi.TYPE_FILE,
                description="User's profile picture.",
            ),
            openapi.Parameter(
                "phone",
                openapi.IN_FORM,
                type=openapi.TYPE_STRING,
                description="User's phone number.",
            ),
            openapi.Parameter(
                "state",
                openapi.IN_FORM,
                type=openapi.TYPE_STRING,
                description="User's state of origin.",
            ),
            openapi.Parameter(
                "school_name",
                openapi.IN_FORM,
                type=openapi.TYPE_STRING,
                description="Candidate's school.",
            ),
            openapi.Parameter(
                "school_type",
                openapi.IN_FORM,
                type=openapi.TYPE_STRING,
                enum=["public", "private"],
                description="Candidate's school type.",
            ),
            openapi.Parameter(
                "current_class",
                openapi.IN_FORM,
                type=openapi.TYPE_STRING,
                enum=["SS1", "SS2", "SS3"],
                description="Candidate's current class.",
            ),
            openapi.Parameter(
                "cowrywise_kid_profile.username",
                openapi.IN_FORM,
                type=openapi.TYPE_STRING,
                description="Candidate's Cowrywise Kids username.",
            ),
            openapi.Parameter(
                "occupation",
                openapi.IN_FORM,
                type=openapi.TYPE_STRING,
                description="Staff's occupation.",
            ),
            api_key,
            bearer_auth,
        ],
        responses={
            200: openapi.Response("Account updated successfully."),
            400: error_response_400,
            401: error_response_401,
            403: error_response_403,
            404: error_response_404,
        },
        tags=["User Management"],
    )
    def patch(self, request, user_id=None):
        """Partially update user and/or profile data."""
        logger.info(
            f"PATCH account_management: user {request.user.id} updating user {user_id}"
        )
        try:
            target_user = self._get_target_user(request, user_id)

            # Extract and validate data
            user_data, profile_data = RequestDataExtractor.extract(request.data)

            if not user_data and not profile_data:
                logger.warning(
                    f"No valid data provided for user {user_id} update. "
                    f"Request data keys: {list(request.data.keys())}"
                )
                return Response(
                    {"error": "No user or profile data provided."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Prepare serializers for validation
            serializers_to_save = []

            # Validate user data if provided
            if user_data:
                from vmlc.serializers import UserSerializer

                user_serializer = UserSerializer(
                    target_user, data=user_data, partial=True
                )
                user_serializer.is_valid(raise_exception=True)
                serializers_to_save.append(user_serializer)

            # Validate profile data if provided
            if profile_data:
                profile, profile_serializer_class = (
                    ProfileManager.get_profile_and_serializer(target_user)
                )
                if profile and profile_serializer_class:
                    profile_serializer = profile_serializer_class(
                        profile, data=profile_data, partial=True
                    )
                    profile_serializer.is_valid(raise_exception=True)
                    serializers_to_save.append(profile_serializer)

            # Save all validated data atomically
            with transaction.atomic():
                for serializer in serializers_to_save:
                    serializer.save()

            # Invalidate caches
            AccountCacheManager.invalidate_user_cache(target_user)

            logger.info(f"Account {target_user.id} updated by {request.user.id}")

            # Return updated profile data
            updated_profile_data = ProfileManager.serialize_profile(target_user)

            return Response(
                {
                    "message": "Account updated successfully.",
                    "profile": updated_profile_data,
                }
            )
        except ValidationError as e:

            def get_first_error(errors):
                if isinstance(errors, dict):
                    key = next(iter(errors))
                    return get_first_error(errors[key])
                elif isinstance(errors, list):
                    return errors[0]
                else:
                    return str(errors)

            error_message = get_first_error(e.detail)
            logger.warning(f"Validation error during account update: {error_message}")
            return Response(
                {"error": str(error_message)}, status=status.HTTP_400_BAD_REQUEST
            )


class BaseInviteView(CreateAPIView):
    """
    Base API view to create a new user invite.
    """

    permission_classes = ActiveManagerPermissions
    serializer_class = None
    profile_type = ""

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user.staff_profile)

    @swagger_auto_schema(
        operation_summary=f"Create {profile_type.title()} User",
        operation_description=(
            f"Creates a new {profile_type.title()} user with the provided details."
            " An invitation email with login credentials will be sent to the user."
        ),
        request_body=(
            staff_registration_request_body
            if profile_type == "staff"
            else candidate_registration_request_body
        ),
        responses={
            201: openapi.Response(f"{profile_type.title()} user created successfully."),
            400: error_response_400,
            401: error_response_401,
            403: error_response_403,
        },
        tags=["User Management"],
        manual_parameters=[api_key, bearer_auth],
    )
    def post(self, request, *args, **kwargs):
        request_data = request.data.copy()
        temp_password = generate_password()
        request_data["password"] = temp_password
        request_data["password2"] = temp_password
        login_url = f"{settings.FRONTEND_LOGIN}"
        serializer = self.get_serializer(data=request_data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        UserVerification.objects.get_or_create(user=serializer.instance.user)
        # Calculate the time delta based on environment
        revoke_delta = timedelta(days=7) if not settings.DEBUG else timedelta(minutes=5)
        time_to_revoke = timezone.now() + revoke_delta

        # Schedule the revocation task to run after 'time_to_revoke'
        revoke_user_invite_task.apply_async(
            args=[serializer.instance.user.id], eta=time_to_revoke
        )

        profile = serializer.instance
        user = profile.user
        profile_msg = ""
        # Dynamically generate the human-readable time string
        if not settings.DEBUG:
            time_to_revoke_str = f"{revoke_delta.days} days"
        else:
            time_to_revoke_str = f"{revoke_delta.seconds // 60} minutes"

        if self.profile_type == "staff":
            profile_msg = (
                "You've been invited to join the Verboheit Mathematics League Competition "
                f"{timezone.now().year} as a staff member. "
            )
        elif self.profile_type == "candidate":
            profile_msg = (
                "Your profile has been created to participate in the "
                "Verboheit Mathematics League Competition. "
            )
        message = (
            f"Hello, {user.first_name}.\n\n"
            f"{profile_msg}"
            f"To accept, log in using the link "
            f"below with this email address and the generated password provided. Please use "
            f"'Forgot Password' to set your own password. If you choose not to accept, simply ignore this message. "
            f"Note that the credentials will expire in {time_to_revoke_str} if you do not log in.\n\n"
            f"Email: {user.email}\n"
            f"Password: {temp_password}\n"
            f"Login: {login_url}\n\n"
            f"Regards,\n"
            f"Management, Verboheit MLC."
        )
        send_mail_task.delay(
            subject="Welcome to Verboheit MLC - Your Account Details",
            message=message,
            recipient_list=[user.email],
        )
        headers = self.get_success_headers({})
        logger.info(
            f"{self.profile_type.title()} profile created successfully with user: {user.id} by user {request.user.id}"
        )
        return Response(
            {"message": f"{self.profile_type.title()} profile created, invite sent."},
            status=status.HTTP_201_CREATED,
            headers=headers,
        )


# @method_decorator(
#     name="post",
#     decorator=swagger_auto_schema(
#         operation_summary="Staff Invite",
#         operation_description="Create a staff profile and send the user an invite.",
#         responses={
#             200: staff_invite_response_schema,
#             401: error_response_401,
#             403: error_response_403,
#         },
#         tags=["Staff"],
#         manual_parameters=[api_key, bearer_auth],
#     ),
# )
class StaffInviteView(BaseInviteView):
    """
    API view to create a new staff member.
    """

    permission_classes = ActiveManagerPermissions
    serializer_class = StaffInviteSerializer
    profile_type = "staff"


@method_decorator(
    name="get",
    decorator=swagger_auto_schema(
        operation_summary="List Users",
        operation_description="List all users. Can be filtered by profile type ('staff' or 'candidate').",
        manual_parameters=[
            openapi.Parameter(
                "profile",
                openapi.IN_QUERY,
                description="Filter by user profile type ('staff' or 'candidate').",
                type=openapi.TYPE_STRING,
            ),
            api_key,
            bearer_auth,
        ],
        responses={
            200: openapi.Response("Paginated list of users."),
            401: error_response_401,
            403: error_response_403,
        },
        tags=["User Management"],
    ),
)
class UserListView(ListAPIView):
    """
    List all users with pagination and optional filtering.
    Can be filtered by `profile` query parameter to 'staff', 'candidate', or 'pre_reg_candidate'.
    Level of detail depends on role.
    """

    permission_classes = ActiveModeratorPermissions
    pagination_class = api_settings.DEFAULT_PAGINATION_CLASS

    def list(self, request, *args, **kwargs):
        user = request.user
        staff = Staff.objects.get(user=user)

        # Get cache version
        version = cache.get("user_list_version", 1)
        # Create a stable cache key based on role, query parameters and version
        query_params_str = str(sorted(request.query_params.items()))
        cache_key = f"user_list_view_{version}_{staff.role}_{query_params_str}"

        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)

        stats_overview = self._get_stats_overview(staff.role)

        queryset_or_list = self.get_queryset()

        # Only call filter_queryset if it's a real QuerySet
        if isinstance(queryset_or_list, QuerySet):
            queryset_or_list = self.filter_queryset(queryset_or_list)

        page = self.paginate_queryset(queryset_or_list)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            response.data["stats_overview"] = stats_overview
            response.data["results"] = response.data.pop("results")
            cache.set(cache_key, response.data, 43200)  # Cache for 12 hours
            return response

        serializer = self.get_serializer(queryset_or_list, many=True)
        response_data = {"stats_overview": stats_overview, "results": serializer.data}
        cache.set(cache_key, response_data, 43200)  # Cache for 12 hour
        return Response(response_data)

    def _get_stats_overview(self, staff_role):
        """Helper to get the stats overview using granular caching."""
        stats_overview = generate_stats_overview_data()

        # Make a copy to avoid modifying the cached objects (since some are shared).
        stats_overview = stats_overview.copy()
        if staff_role not in [Staff.Roles.MANAGER, Staff.Roles.SUPERADMIN]:
            stats_overview.pop("staff", None)

        return stats_overview

    def get_serializer_class(self):
        requested_profile = self.request.query_params.get("profile")
        if requested_profile == "candidate":
            return CandidateListSerializer
        if requested_profile == "staff":
            return StaffListSerializer
        if requested_profile in ["pre_reg_candidate", "pre_reg_staff"]:
            return PreRegUserSerializer
        return UserProfileListSerializer

    def get_serializer_kwargs(self):
        kwargs = super().get_serializer_kwargs()
        requested_profile = self.request.query_params.get("profile")
        if requested_profile == "pre_reg_candidate":
            kwargs["interest_type"] = "candidate"
        if requested_profile == "pre_reg_staff":
            kwargs["interest_type"] = "volunteer"
        return kwargs

    def get_queryset(self):
        """Returns a filtered queryset or combined list of users"""

        requested_profile = self.request.query_params.get("profile")
        logger.info(
            f"UserListView: request from user {self.request.user.id} with query params: {self.request.query_params}"
        )
        if requested_profile == "candidate":
            queryset = Candidate.objects.select_related("user").order_by("-created_at")
            return filter_candidates(queryset, self.request.query_params)
        if requested_profile == "staff":
            queryset = Staff.objects.select_related("user").order_by("-created_at")
            return filter_staffs(queryset, self.request.query_params)
        if requested_profile in ["pre_reg_candidate", "pre_reg_staff"]:
            queryset = PreRegUser.objects.order_by("-created_at")
            return filter_pre_reg_users(queryset, self.request.query_params)

        # Combined view: fully registered and pre-registered users
        users_qs = User.objects.prefetch_related(
            "staff_profile", "candidate_profile", "verification"
        ).order_by("-date_joined")
        users_qs = filter_users(users_qs, self.request.query_params)

        # Pre-registered users should also be filtered by common params like search
        pre_reg_qs = PreRegUser.objects.order_by("-created_at")
        search = self.request.query_params.get("search")
        if search:
            pre_reg_qs = pre_reg_qs.filter(
                Q(full_name__icontains=search) | Q(email__icontains=search)
            )

        # Don't show pre-registered users if filtering by role/school since they don't have them
        role = self.request.query_params.get("role")
        school = self.request.query_params.get("school_name")
        if role or school:
            return users_qs

        # Merge and sort by join/creation date
        combined = sorted(
            itertools.chain(users_qs, pre_reg_qs),
            key=lambda x: x.date_joined if isinstance(x, User) else x.created_at,
            reverse=True,
        )
        return combined


@method_decorator(
    name="get",
    decorator=swagger_auto_schema(
        operation_summary="Get User Profile",
        operation_description="Retrieve details for a specific staff or candidate profile.",
        responses={
            200: openapi.Response(
                "Profile details for a staff or candidate member. The schema depends on the user profile.",
                schema=UserProfileDetailSerializer,
            ),
            401: error_response_401,
            403: error_response_403,
            404: error_response_404,
        },
        tags=["User Management"],
        manual_parameters=[api_key, bearer_auth],
    ),
)
@method_decorator(
    name="patch",
    decorator=swagger_auto_schema(
        operation_summary="Update User Profile",
        operation_description="Partially update a staff or candidate profile. The available fields depend on the profile type.",
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            properties={
                "school_name": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="Candidate's school. (Candidate only)",
                ),
                "occupation": openapi.Schema(
                    type=openapi.TYPE_STRING,
                    description="Staff's occupation. (Staff only)",
                ),
                "role": openapi.Schema(
                    type=openapi.TYPE_STRING, description="User role within the system."
                ),
                "is_active": openapi.Schema(
                    type=openapi.TYPE_BOOLEAN, description="Set profile active status."
                ),
            },
        ),
        responses={
            200: openapi.Response(
                "Profile updated successfully.", schema=UserProfileDetailSerializer
            ),
            400: error_response_400,
            401: error_response_401,
            403: error_response_403,
            404: error_response_404,
        },
        tags=["User Management"],
        manual_parameters=[api_key, bearer_auth],
    ),
)
@method_decorator(
    name="delete",
    decorator=swagger_auto_schema(
        operation_summary="Deactivate User Profile",
        operation_description="Deactivates a staff or candidate profile by setting `is_active` to false (soft delete).",
        responses={
            204: openapi.Response("Profile deactivated successfully."),
            401: error_response_401,
            403: error_response_403,
            404: error_response_404,
        },
        tags=["User Management"],
        manual_parameters=[api_key, bearer_auth],
    ),
)
class UserDetailView(RetrieveUpdateDestroyAPIView):
    """
    Retrieve, update, or deactivate a specific staff or candidate member.

    - GET: Retrieve profile details.
    - PATCH: Update profile.
    - DELETE: Soft delete the profile (marks as inactive).

    Access to staff profiles is restricted to Managers and Superadmins.
    """

    serializer_class = UserProfileDetailSerializer
    permission_classes = ActiveAdminPermissions + [IsManagerForStaffDetail]
    http_method_names = ["get", "patch", "delete"]
    profile = ""

    def initial(self, request, *args, **kwargs):
        """
        Dynamically set lookup_url_kwarg and profile based on URL.
        """
        super().initial(request, *args, **kwargs)
        if "staff_id" in self.kwargs:
            self.lookup_url_kwarg = "staff_id"
            self.profile = "staff"
        elif "candidate_id" in self.kwargs:
            self.lookup_url_kwarg = "candidate_id"
            self.profile = "candidate"

    def retrieve(self, request, *args, **kwargs):
        """Get candidate/staff profile detail"""
        profile_id = self.kwargs.get(self.lookup_url_kwarg)
        cache_key = f"{self.profile}_profile_{profile_id}"

        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)

        instance = self.get_object()
        serializer = self.get_serializer(instance)
        data = serializer.data
        cache.set(cache_key, data, 3600)  # Cache for 1 hour
        return Response(data)

    def perform_update(self, serializer):
        """Save updates to candidate/staff"""
        target_profile = serializer.instance
        request_user = self.request.user
        logger.info(
            "%s profile %s update request by user %s",
            self.profile.title(),
            target_profile.pk,
            request_user.id,
        )
        serializer.save()
        cache.delete(f"{self.profile}_profile_{target_profile.pk}")
        cache.delete(f"{self.profile}_dashboard_data_{target_profile.pk}")

    def perform_destroy(self, instance):
        """Deactivates a candidate/staff profile (soft delete)"""
        target_profile = instance
        request_user = self.request.user
        logger.info(
            "%s profile %s soft-delete request by user %s",
            self.profile.title(),
            target_profile.pk,
            request_user.id,
        )
        instance.is_active = False
        instance.save()
        cache.delete(f"{self.profile}_profile_{target_profile.pk}")
        cache.delete(f"{self.profile}_dashboard_data_{target_profile.pk}")

    def get_queryset(self):
        if self.profile == "candidate":
            return (
                Candidate.objects.select_related("user", "user__verification")
                .prefetch_related("results__exam", "results__score_submitted_by__user")
                .all()
            )
        if self.profile == "staff":
            return Staff.objects.select_related("user", "user__verification").all()
        return User.objects.none()


class BulkNotificationView(APIView):
    """
    Send bulk notification/email/SMS to selected users.
    """
    permission_classes = ActiveManagerPermissions

    def post(self, request):
        from comms.models import Notification, Broadcast
        from comms.signals import notifications_created
        from comms.tasks import send_bulk_sms_task
        from vmlc.serializers.comms import BulkNotificationSerializer
        import base64
        import uuid
        import os
        from django.conf import settings

        serializer = BulkNotificationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data

        user_ids = validated_data["user_ids"]
        message = validated_data["message"]
        medium = validated_data.get("medium", "email")
        subject = validated_data.get("subject", "")

        # Get user profiles
        users = User.objects.filter(id__in=user_ids)
        target_count = users.count()

        if target_count == 0:
            return Response(
                {"error": "No valid users found."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Map medium string to Broadcast mediums list
        medium_map = {
            "email": [Broadcast.Medium.EMAIL],
            "sms": [Broadcast.Medium.SMS],
            "both": [Broadcast.Medium.EMAIL, Broadcast.Medium.SMS],
        }

        # Create a Broadcast record to track this batch
        broadcast = Broadcast.objects.create(
            created_by=request.user.staff_profile,
            subject=subject or "Bulk Notification",
            message=message,
            mediums=medium_map.get(medium, [Broadcast.Medium.EMAIL]),
            target_roles={},
            total_recipients=target_count,
            status=Broadcast.Status.IN_PROGRESS,
        )

        # Determine delivery options
        send_email = medium in ["email", "both"]
        send_sms = medium in ["sms", "both"]

        # Handle attached image
        image_path = None
        image_raw = validated_data.get("image_base64", "")
        image_filename = validated_data.get("image_name", "")
        if image_raw and send_email:
            try:
                image_data = base64.b64decode(image_raw)
                ext = os.path.splitext(image_filename)[1] or ".png"
                safe_name = f"{uuid.uuid4()}{ext}"
                upload_dir = os.path.join(settings.MEDIA_ROOT, "broadcast_images")
                os.makedirs(upload_dir, exist_ok=True)
                image_path = os.path.join(upload_dir, safe_name)
                with open(image_path, "wb") as f:
                    f.write(image_data)
            except Exception as e:
                logger.warning("Failed to save attached image: %s", e)

        # Collect phone numbers for SMS
        sms_recipients = []
        if send_sms:
            for user in users:
                if user.phone:
                    sms_recipients.append(user.phone)

        # We create a Notification record for each user.
        # This provides history and dashboard visibility.
        metadata = {"broadcast_id": broadcast.id}
        if image_path:
            metadata["image_path"] = image_path
            metadata["image_name"] = image_filename or "image.png"
        if send_email:
            metadata["send_email"] = True
        if send_sms:
            metadata["send_sms"] = True

        notifications = [
            Notification(
                recipient=user,
                subject=subject,
                message=message,
                type=Notification.Type.INFO,
                metadata=metadata,
            )
            for user in users
        ]

        try:
            with transaction.atomic():
                created_notifications = Notification.objects.bulk_create(notifications)

                # Trigger background delivery (WebSocket and/or Email)
                notifications_created.send(
                    sender=self.__class__,
                    notifications=created_notifications
                )

                # Send SMS if needed
                if send_sms and sms_recipients:
                    sms_message = f"{subject}:\n\n{message}" if subject else message
                    send_bulk_sms_task.delay(
                        body=sms_message,
                        recipients=sms_recipients,
                        broadcast_log_id=None,
                        broadcast_id=broadcast.id,
                    )

        except Exception as e:
            logger.error(f"Failed to create bulk notifications: {e}")
            broadcast.status = Broadcast.Status.FAILED
            broadcast.save(update_fields=["status"])
            return Response(
                {"error": f"Failed to create notifications: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        delivery_methods = []
        if send_email:
            delivery_methods.append("email")
        if send_sms:
            delivery_methods.append("SMS")

        return Response({
            "message": f"{', '.join(delivery_methods)} sent to {target_count} users.",
            "broadcast_id": broadcast.id,
        })


class ResetUserPasswordView(APIView):
    """
    Reset a user's password and send them a temporary password via email.
    """
    permission_classes = ActiveManagerPermissions

    def post(self, request):
        from identity.models import User
        from vmlc.utils.auth import generate_password
        from comms.tasks import send_mail_task
        from django.conf import settings

        user_id = request.data.get("user_id")

        if not user_id:
            return Response(
                {"error": "user_id is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response(
                {"error": "User not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Generate temporary password
        temp_password = generate_password()
        user.set_password(temp_password)
        user.save(update_fields=["password"])

        # Send email with temporary password
        subject = "Password Reset - Verbohit MLC"
        message = (
            f"Hello {user.first_name},\n\n"
            f"Your password has been reset by an administrator.\n\n"
            f"Your temporary password is: {temp_password}\n\n"
            f"Please log in and change your password immediately.\n\n"
            f"Login URL: {settings.FRONTEND_LOGIN}\n\n"
            f"Regards,\n"
            f"Management, Verbohit MLC."
        )

        try:
            send_mail_task.delay(
                subject=subject,
                message=message,
                recipient_list=[user.email],
            )
        except Exception as e:
            logger.error(f"Failed to send password reset email: {e}")
            return Response(
                {"error": "Password reset but failed to send email. Please contact IT."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        return Response({
            "message": f"Password reset for {user.email}. Temporary password sent via email.",
        })


class UserActivityLogView(ListAPIView):
    """
    Retrieve activity log for a specific user.
    """
    permission_classes = ActiveModeratorPermissions
    serializer_class = None

    def list(self, request, *args, **kwargs):
        from vmlc.models import Event
        from identity.models import User

        user_id = request.query_params.get("user_id")

        if not user_id:
            return Response(
                {"error": "user_id query parameter is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response(
                {"error": "User not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Get events for this user
        events = Event.objects.filter(actor=user).order_by("-timestamp")[:50]

        return Response({
            "user": {
                "id": user.id,
                "email": user.email,
                "first_name": user.first_name,
                "last_name": user.last_name,
            },
            "activities": [
                {
                    "event_name": e.event_name,
                    "timestamp": e.timestamp.isoformat(),
                    "metadata": e.metadata,
                }
                for e in events
            ]
        })


class BulkStaffImportView(APIView):
    """
    Bulk import staff from CSV/Excel file.
    """
    permission_classes = ActiveManagerPermissions
    parser_classes = [MultiPartParser]

    def post(self, request):
        import io
        from openpyxl import load_workbook

        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Determine file type and load
        filename = file.name.lower()
        rows = []

        try:
            if filename.endswith(".csv"):
                # Read CSV
                content = file.read().decode("utf-8")
                import csv
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif filename.endswith((".xlsx", ".xls")):
                # Read Excel
                wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                # Skip header row, read all data rows
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):  # Skip empty rows
                        row_dict = dict(zip(headers, row))
                        rows.append(row_dict)
                wb.close()
            else:
                return Response(
                    {"error": "Invalid file format. Only CSV and Excel files are supported."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            logger.error(f"Failed to parse file: {e}")
            return Response(
                {"error": f"Failed to parse file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not rows:
            return Response(
                {"error": "No data found in file."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Expected columns
        required_fields = ["email", "first_name", "last_name"]
        optional_fields = ["phone", "state", "role", "occupation"]
        all_fields = required_fields + optional_fields

        # Validate columns from first row
        first_row = rows[0]
        missing_required = [f for f in required_fields if f not in first_row]
        if missing_required:
            return Response(
                {"error": f"Missing required columns: {', '.join(missing_required)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Process each row
        results = {
            "success": [],
            "errors": [],
            "created": 0,
            "failed": 0,
        }

        VALID_ROLES = ["volunteer", "moderator", "admin", "manager", "superadmin"]
        VALID_STATES = ["lagos", "ogun", "rivers", "abuja"]

        for idx, row in enumerate(rows, start=2):  # Start at 2 to account for header
            email = row.get("email", "").strip().lower()
            first_name = row.get("first_name", "").strip()
            last_name = row.get("last_name", "").strip()
            phone = row.get("phone", "").strip()
            state = row.get("state", "").strip().lower()
            role = row.get("role", "volunteer").strip().lower()
            occupation = row.get("occupation", "").strip()

            # Validate required fields
            if not email:
                results["errors"].append({
                    "row": idx,
                    "error": "Email is required",
                })
                results["failed"] += 1
                continue

            if not first_name:
                results["errors"].append({
                    "row": idx,
                    "error": "First name is required",
                })
                results["failed"] += 1
                continue

            if not last_name:
                results["errors"].append({
                    "row": idx,
                    "error": "Last name is required",
                })
                results["failed"] += 1
                continue

            # Validate role
            if role and role not in VALID_ROLES:
                results["errors"].append({
                    "row": idx,
                    "error": f"Invalid role '{role}'. Valid roles: {', '.join(VALID_ROLES)}",
                })
                results["failed"] += 1
                continue

            # Validate state
            if state and state not in VALID_STATES:
                results["errors"].append({
                    "row": idx,
                    "error": f"Invalid state '{state}'. Valid states: {', '.join(VALID_STATES)}",
                })
                results["failed"] += 1
                continue

            # Check if user exists
            try:
                if User.objects.filter(email__iexact=email).exists():
                    results["errors"].append({
                        "row": idx,
                        "error": f"Email '{email}' already exists",
                    })
                    results["failed"] += 1
                    continue
            except Exception as e:
                results["errors"].append({
                    "row": idx,
                    "error": f"Database error: {str(e)}",
                })
                results["failed"] += 1
                continue

            # Generate password and create user
            try:
                temp_password = generate_password()
                user = User.objects.create_user(
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    phone=phone or None,
                    state=state or None,
                    password=temp_password,
                )

                staff = Staff.objects.create(
                    user=user,
                    role=role or "volunteer",
                    occupation=occupation or "",
                    created_by=request.user.staff_profile,
                )

                # Send invitation email
                try:
                    subject = "Welcome to Verbohit MLC"
                    message = (
                        f"Hello {first_name},\n\n"
                        f"You have been invited to join Verbohit MLC as a {role or 'volunteer'}.\n\n"
                        f"Your login credentials:\n"
                        f"Email: {email}\n"
                        f"Password: {temp_password}\n\n"
                        f"Please log in and change your password immediately.\n\n"
                        f"Login URL: {settings.FRONTEND_LOGIN}\n\n"
                        f"Regards,\n"
                        f"Management, Verbohit MLC."
                    )
                    send_mail_task.delay(
                        subject=subject,
                        message=message,
                        recipient_list=[email],
                    )
                except Exception as e:
                    logger.warning(f"Failed to send invitation email for {email}: {e}")

                results["success"].append({
                    "row": idx,
                    "email": email,
                    "staff_id": str(staff.pk),
                })
                results["created"] += 1

            except Exception as e:
                results["errors"].append({
                    "row": idx,
                    "error": f"Failed to create staff: {str(e)}",
                })
                results["failed"] += 1
                continue

        return Response(results)


class BulkCandidateImportView(APIView):
    """
    Bulk import candidates from CSV/Excel file.
    """
    permission_classes = ActiveManagerPermissions
    parser_classes = [MultiPartParser]

    def post(self, request):
        import io
        from openpyxl import load_workbook

        file = request.FILES.get("file")
        if not file:
            return Response(
                {"error": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Determine file type and load
        filename = file.name.lower()
        rows = []

        try:
            if filename.endswith(".csv"):
                content = file.read().decode("utf-8")
                import csv
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif filename.endswith((".xlsx", ".xls")):
                wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        row_dict = dict(zip(headers, row))
                        rows.append(row_dict)
                wb.close()
            else:
                return Response(
                    {"error": "Invalid file format. Only CSV and Excel files are supported."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            logger.error(f"Failed to parse file: {e}")
            return Response(
                {"error": f"Failed to parse file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not rows:
            return Response(
                {"error": "No data found in file."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Expected columns
        required_fields = ["email", "first_name", "last_name"]
        missing_required = [f for f in required_fields if f not in rows[0]]
        if missing_required:
            return Response(
                {"error": f"Missing required columns: {', '.join(missing_required)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        results = {
            "success": [],
            "errors": [],
            "created": 0,
            "failed": 0,
        }

        VALID_CLASSES = ["SS1", "SS2", "SS3"]
        VALID_SCHOOL_TYPES = ["public", "private"]
        VALID_STATES = ["lagos", "ogun", "rivers", "abuja"]
        VALID_ROLES = ["screening", "league", "final", "winner"]

        for idx, row in enumerate(rows, start=2):
            email = row.get("email", "").strip().lower()
            first_name = row.get("first_name", "").strip()
            last_name = row.get("last_name", "").strip()
            phone = row.get("phone", "").strip()
            state = row.get("state", "").strip().lower()
            school_name = row.get("school_name", "").strip()
            school_type = row.get("school_type", "").strip().lower()
            current_class = row.get("current_class", "").strip().upper()

            # Validate required fields
            if not email:
                results["errors"].append({"row": idx, "error": "Email is required"})
                results["failed"] += 1
                continue
            if not first_name:
                results["errors"].append({"row": idx, "error": "First name is required"})
                results["failed"] += 1
                continue
            if not last_name:
                results["errors"].append({"row": idx, "error": "Last name is required"})
                results["failed"] += 1
                continue

            # Validate optional fields
            if current_class and current_class not in VALID_CLASSES:
                results["errors"].append({
                    "row": idx,
                    "error": f"Invalid class '{current_class}'. Valid: {', '.join(VALID_CLASSES)}",
                })
                results["failed"] += 1
                continue

            if school_type and school_type not in VALID_SCHOOL_TYPES:
                results["errors"].append({
                    "row": idx,
                    "error": f"Invalid school_type '{school_type}'. Valid: {', '.join(VALID_SCHOOL_TYPES)}",
                })
                results["failed"] += 1
                continue

            if state and state not in VALID_STATES:
                results["errors"].append({
                    "row": idx,
                    "error": f"Invalid state '{state}'. Valid: {', '.join(VALID_STATES)}",
                })
                results["failed"] += 1
                continue

            # Check if user exists
            if User.objects.filter(email__iexact=email).exists():
                results["errors"].append({"row": idx, "error": f"Email '{email}' already exists"})
                results["failed"] += 1
                continue

            # Create user and candidate
            try:
                temp_password = generate_password()
                user = User.objects.create_user(
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    phone=phone or None,
                    state=state or None,
                    password=temp_password,
                )

                candidate = Candidate.objects.create(
                    user=user,
                    school_name=school_name or "",
                    school_type=school_type or "public",
                    current_class=current_class or "SS1",
                    role=Candidate.Roles.SCREENING,
                    created_by=request.user.staff_profile,
                )

                # Send invitation email
                try:
                    subject = "Welcome to Verbohit MLC"
                    message = (
                        f"Hello {first_name},\n\n"
                        f"You have been invited to participate in Verbohit MLC.\n\n"
                        f"Your login credentials:\n"
                        f"Email: {email}\n"
                        f"Password: {temp_password}\n\n"
                        f"Please log in and change your password immediately.\n\n"
                        f"Login URL: {settings.FRONTEND_LOGIN}\n\n"
                        f"Regards,\n"
                        f"Management, Verbohit MLC."
                    )
                    send_mail_task.delay(
                        subject=subject,
                        message=message,
                        recipient_list=[email],
                    )
                except Exception as e:
                    logger.warning(f"Failed to send invitation email for {email}: {e}")

                results["success"].append({
                    "row": idx,
                    "email": email,
                    "candidate_id": str(candidate.pk),
                })
                results["created"] += 1

            except Exception as e:
                results["errors"].append({
                    "row": idx,
                    "error": f"Failed to create candidate: {str(e)}",
                })
                results["failed"] += 1
                continue

        return Response(results)
