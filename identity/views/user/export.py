import io
import logging
from datetime import datetime
from collections import defaultdict

from django.http import HttpResponse
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from identity.models import Staff, Candidate
from identity.permissions import ActiveManagerPermissions
from vmlc.utils.query_filters import filter_staffs, filter_candidates
from competition.models import Competition, Stage, StageExam, Enrollment
from vmlc.models import CandidateExamResult, Exam

logger = logging.getLogger(__name__)

CANDIDATE_COLUMNS = {
    "sn": {"label": "S/N", "key": "sn"},
    "full_name": {"label": "Full Name", "key": "full_name"},
    "email": {"label": "Email", "key": "email"},
    "phone": {"label": "Phone", "key": "phone"},
    "school_name": {"label": "School Name", "key": "school_name"},
    "school_type": {"label": "School Type", "key": "school_type"},
    "current_class": {"label": "Current Class", "key": "current_class"},
    "state": {"label": "State", "key": "state"},
    "role": {"label": "Role", "key": "role"},
    "date_joined": {"label": "Date Joined", "key": "date_joined"},
    "status": {"label": "Status", "key": "status"},
}

STAFF_COLUMNS = {
    "sn": {"label": "S/N", "key": "sn"},
    "full_name": {"label": "Full Name", "key": "full_name"},
    "email": {"label": "Email", "key": "email"},
    "phone": {"label": "Phone", "key": "phone"},
    "occupation": {"label": "Occupation", "key": "occupation"},
    "role": {"label": "Role", "key": "role"},
    "date_joined": {"label": "Date Joined", "key": "date_joined"},
    "status": {"label": "Status", "key": "status"},
}

ALL_CANDIDATE_COLUMNS = list(CANDIDATE_COLUMNS.keys())
ALL_STAFF_COLUMNS = list(STAFF_COLUMNS.keys())


class UserExportView(APIView):
    """
    Export user data (staff or candidates) to Excel format.
    Inherits filters and sorting from the list view.
    Supports column selection via 'columns' query parameter (comma-separated).
    """
    permission_classes = ActiveManagerPermissions

    def get(self, request, *args, **kwargs):
        profile_type = request.query_params.get("profile")
        columns_param = request.query_params.get("columns", "")
        include_exam_data = request.query_params.get("include_exam_data", "").lower() == "true"

        if profile_type == "candidate":
            queryset = Candidate.objects.select_related("user").order_by("-created_at")
            queryset = filter_candidates(queryset, request.query_params)
            filename = f"candidates_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            if columns_param:
                columns = [c.strip() for c in columns_param.split(",") if c.strip()]
                columns = [c for c in columns if c in ALL_CANDIDATE_COLUMNS]
            else:
                columns = ALL_CANDIDATE_COLUMNS

            exam_columns = []
            exam_data_map = {}
            if include_exam_data:
                exam_columns, exam_data_map = self._prepare_exam_data(queryset)

            return self._export_candidates(queryset, filename, columns, exam_columns, exam_data_map)

        elif profile_type == "staff":
            queryset = Staff.objects.select_related("user").order_by("-created_at")
            queryset = filter_staffs(queryset, request.query_params)
            filename = f"staff_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            if columns_param:
                columns = [c.strip() for c in columns_param.split(",") if c.strip()]
                columns = [c for c in columns if c in ALL_STAFF_COLUMNS]
            else:
                columns = ALL_STAFF_COLUMNS

            return self._export_staff(queryset, filename, columns)

        else:
            return HttpResponse("Invalid profile type. Must be 'candidate' or 'staff'.", status=400)

    def _prepare_exam_data(self, queryset):
        """
        Prepare exam columns and data map for candidates.
        Returns (exam_columns, exam_data_map) where:
        - exam_columns: list of dicts with 'key' and 'label' for each exam
        - exam_data_map: dict of {candidate_id: {exam_key: score}}
        """
        # Get active competition
        try:
            competition = Competition.objects.filter(status=Competition.Status.ACTIVE).first()
            if not competition:
                competition = Competition.objects.order_by("-edition").first()
        except Competition.DoesNotExist:
            return [], {}

        # Get all stage exams for the competition, ordered by stage order then round
        stage_exams = (
            StageExam.objects
            .filter(competition_stage__competition=competition)
            .select_related("competition_stage", "exam")
            .order_by("competition_stage__order", "round")
        )

        # Build exam columns
        exam_columns = []
        exam_key_map = {}  # exam_id -> key

        for slot in stage_exams:
            try:
                exam = slot.exam
            except Exam.DoesNotExist:
                continue

            stage_type = slot.competition_stage.type

            if slot.round:
                key = f"exam_{stage_type}_r{slot.round}"
                label = f"{slot.competition_stage.get_type_display()} R{slot.round}"
            else:
                key = f"exam_{stage_type}"
                label = slot.competition_stage.get_type_display()

            exam_columns.append({"key": key, "label": label})
            exam_key_map[exam.id] = key

        # Add last_stage column
        exam_columns.append({"key": "last_stage", "label": "Last Stage"})

        # Build exam data map: {candidate_id: {exam_key: score}}
        exam_data_map = defaultdict(dict)

        # Get all results for candidates in queryset
        candidate_ids = [c.pk for c in queryset]
        results = (
            CandidateExamResult.objects
            .filter(candidate_id__in=candidate_ids, exam_id__in=exam_key_map.keys())
            .select_related("exam")
        )

        for result in results:
            key = exam_key_map.get(result.exam_id)
            if key:
                exam_data_map[result.candidate_id][key] = float(result.score)

        # Get last stage for each candidate from enrollment
        enrollments = (
            Enrollment.objects
            .filter(candidate_id__in=candidate_ids, competition=competition)
            .select_related("current_stage")
        )

        for enrollment in enrollments:
            if enrollment.current_stage:
                exam_data_map[enrollment.candidate_id]["last_stage"] = enrollment.current_stage.get_type_display()
            else:
                exam_data_map[enrollment.candidate_id]["last_stage"] = "N/A"

        return exam_columns, exam_data_map

    def _export_candidates(self, queryset, filename, columns, exam_columns=None, exam_data_map=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"

        # Build headers
        headers = [CANDIDATE_COLUMNS.get(c, {}).get("label", c) for c in columns]
        if exam_columns:
            headers.extend([col["label"] for col in exam_columns])

        self._setup_sheet(ws, headers)

        for index, candidate in enumerate(queryset, start=1):
            user = candidate.user
            row_data = []

            for col in columns:
                if col == "sn":
                    row_data.append(index)
                elif col == "full_name":
                    row_data.append(f"{user.first_name} {user.last_name}")
                elif col == "email":
                    row_data.append(user.email)
                elif col == "phone":
                    row_data.append(user.phone or "N/A")
                elif col == "school_name":
                    row_data.append(candidate.school_name)
                elif col == "school_type":
                    row_data.append(candidate.get_school_type_display() if hasattr(candidate, 'get_school_type_display') else candidate.school_type)
                elif col == "current_class":
                    row_data.append(candidate.current_class or "N/A")
                elif col == "state":
                    row_data.append(user.state or "N/A")
                elif col == "role":
                    row_data.append(candidate.get_role_display())
                elif col == "date_joined":
                    row_data.append(user.date_joined.strftime("%Y-%m-%d %H:%M:%S") if user.date_joined else "N/A")
                elif col == "status":
                    row_data.append(candidate.status)
                else:
                    row_data.append("")

            # Add exam data
            if exam_columns:
                candidate_exam_data = exam_data_map.get(candidate.pk, {})
                for col in exam_columns:
                    if col["key"] == "last_stage":
                        row_data.append(candidate_exam_data.get("last_stage", "N/A"))
                    else:
                        score = candidate_exam_data.get(col["key"])
                        row_data.append(score if score is not None else "N/A")

            ws.append(row_data)

        return self._get_response(wb, filename)

    def _export_staff(self, queryset, filename, columns):
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff"

        headers = [STAFF_COLUMNS.get(c, {}).get("label", c) for c in columns]
        self._setup_sheet(ws, headers)

        for index, staff in enumerate(queryset, start=1):
            user = staff.user
            row_data = []

            for col in columns:
                if col == "sn":
                    row_data.append(index)
                elif col == "full_name":
                    row_data.append(f"{user.first_name} {user.last_name}")
                elif col == "email":
                    row_data.append(user.email)
                elif col == "phone":
                    row_data.append(user.phone or "N/A")
                elif col == "occupation":
                    row_data.append(staff.occupation or "N/A")
                elif col == "role":
                    row_data.append(staff.get_role_display())
                elif col == "date_joined":
                    row_data.append(user.date_joined.strftime("%Y-%m-%d %H:%M:%S") if user.date_joined else "N/A")
                elif col == "status":
                    row_data.append(staff.status)
                else:
                    row_data.append("")

            ws.append(row_data)

        return self._get_response(wb, filename)

    def _setup_sheet(self, ws, headers):
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3E4095", end_color="3E4095", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

    def _get_response(self, wb, filename):
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"
        return response
