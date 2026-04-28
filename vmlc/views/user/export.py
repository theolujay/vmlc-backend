import io
import logging
from datetime import datetime

from django.http import HttpResponse
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from identity.models import Staff, Candidate
from identity.permissions import ActiveManagerPermissions
from vmlc.utils.query_filters import filter_staffs, filter_candidates

logger = logging.getLogger(__name__)

class UserExportView(APIView):
    """
    Export user data (staff or candidates) to Excel format.
    Inherits filters and sorting from the list view.
    """
    permission_classes = ActiveManagerPermissions

    def get(self, request, *args, **kwargs):
        profile_type = request.query_params.get("profile")

        if profile_type == "candidate":
            queryset = Candidate.objects.select_related("user").order_by("-created_at")
            queryset = filter_candidates(queryset, request.query_params)
            filename = f"candidates_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return self._export_candidates(queryset, filename)

        elif profile_type == "staff":
            queryset = Staff.objects.select_related("user").order_by("-created_at")
            queryset = filter_staffs(queryset, request.query_params)
            filename = f"staff_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return self._export_staff(queryset, filename)

        else:
            return HttpResponse("Invalid profile type. Must be 'candidate' or 'staff'.", status=400)

    def _export_candidates(self, queryset, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"

        headers = [
            "S/N", "Full Name", "Email", "Phone", "School Name", "School Type",
            "Current Class", "State", "Role", "Date Joined", "Status"
        ]

        self._setup_sheet(ws, headers)

        for index, candidate in enumerate(queryset, start=1):
            user = candidate.user
            ws.append([
                index,
                f"{user.first_name} {user.last_name}",
                user.email,
                user.phone or "N/A",
                candidate.school_name,
                candidate.get_school_type_display() if hasattr(candidate, 'get_school_type_display') else candidate.school_type,
                candidate.current_class or "N/A",
                user.state or "N/A",
                candidate.get_role_display(),
                user.date_joined.strftime("%Y-%m-%d %H:%M:%S") if user.date_joined else "N/A",
                candidate.status
            ])

        return self._get_response(wb, filename)

    def _export_staff(self, queryset, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff"

        headers = [
            "S/N", "Full Name", "Email", "Phone", "Occupation", "Role", "Date Joined", "Status"
        ]

        self._setup_sheet(ws, headers)

        for index, staff in enumerate(queryset, start=1):
            user = staff.user
            ws.append([
                index,
                f"{user.first_name} {user.last_name}",
                user.email,
                user.phone or "N/A",
                staff.occupation or "N/A",
                staff.get_role_display(),
                user.date_joined.strftime("%Y-%m-%d %H:%M:%S") if user.date_joined else "N/A",
                staff.status
            ])

        return self._get_response(wb, filename)

    def _setup_sheet(self, ws, headers):
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3E4095", end_color="3E4095", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

    def _get_response(self, wb, filename):
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"
        return response
